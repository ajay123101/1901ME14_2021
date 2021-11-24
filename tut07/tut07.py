import csv
import os

import openpyxl
roll_data = {}
sub_ltp = {}
sub_data = {}
stud_info = {}
def get_stud_info():
    with open('studentinfo.csv', 'r') as temp:
        read = csv.DictReader(temp)
        for row in read:
            val = row['Roll No']
            if val not in stud_info:
                stud_info[val] = {}
            stud_info[val]['name'] = row['Name']
            stud_info[val]['email'] = row['email']
            stud_info[val]['aemail'] = row['aemail']
            stud_info[val]['contact'] = row['contact']
        # print(stud_info)

def get_sub_data():
    with open('course_feedback_submitted_by_students.csv', 'r') as temp:
        read = csv.DictReader(temp)
        for row in read:
            roll = row['stud_roll']
            if roll not in sub_data:
                sub_data[roll] = {}
            if row['course_code'] not in sub_data[roll]:
                sub_data[roll][row['course_code']] = 1
            else:
                sub_data[roll][row['course_code']] += 1
        # print(sub_data)
def get_roll_data():
    with open('course_registered_by_all_students.csv', 'r') as temp:
        read = csv.DictReader(temp)
        for row in read:
            temp = []
            val = row['rollno']
            if val not in roll_data:
                roll_data[val] = temp
            temp2 = {}
            temp2[row['subno']] = [row['register_sem'], row['schedule_sem']]
            roll_data[val].append(temp2)
def get_sub_ltp():
    with open('course_master_dont_open_in_excel.csv', 'r') as temp:
        read = csv.DictReader(temp)
        for row in read:
            val = row['subno']
            n = 0
            item = row['ltp'].split('-')
            for i in item:
                if i != '0':
                    n += 1
            sub_ltp[val] = n

        # print(sub_ltp)
def feedback_not_submitted():

    ltp_mapping_feedback_type = {1: 'lecture', 2: 'tutorial', 3: 'practical'}
    n = 0
    output_file_name = "course_feedback_remaining.xlsx"
    # wb = openpyxl.load_workbook(output_file_name)
    wb = openpyxl.load_workbook(filename=output_file_name) # Older method was  .get_sheet_by_name('Sheet1')
    sheet = wb.active
    # feed back filled by students
    get_stud_info()
    get_sub_data()
    get_sub_ltp()
    # roll no and their respected subject taken
    get_roll_data()
    for roll in roll_data:
        print(n)
        n += 1
        # print(roll)
        # print(roll_data[roll])
        try:
            for j in roll_data[roll]:
                sub = list(j.keys())[0]
                # print(sub)
                if roll in sub_data:
                    if sub in sub_data[roll]:
                        if sub_data[roll][sub] == sub_ltp[sub]:
                            pass
                        else:
                            ls = (
                                roll, j[sub][0], j[sub][1], sub, stud_info[roll]['name'], stud_info[roll]['email'],
                                stud_info[roll]['aemail'], stud_info[roll]['contact'])
                            row = sheet.max_row + 1

                            for col, entry in enumerate(ls, start=1):
                                sheet.cell(row=row, column=col, value=entry)

                    else:
                        ls = (roll, j[sub][0], j[sub][1], sub, stud_info[roll]['name'], stud_info[roll]['email'],
                              stud_info[roll]['aemail'], stud_info[roll]['contact'])
                        row = sheet.max_row + 1
                        for col, entry in enumerate(ls, start=1):
                            sheet.cell(row=row, column=col, value=entry)

                else:
                    # print(j)
                    ls = (roll, j[sub][0], j[sub][1], sub, stud_info[roll]['name'], stud_info[roll]['email'],
                          stud_info[roll]['aemail'], stud_info[roll]['contact'])
                    row = sheet.max_row + 1
                    for col, entry in enumerate(ls, start=1):
                        sheet.cell(row=row, column=col, value=entry)
                wb.save(output_file_name)
        except:
            pass

if not os.path.isfile("course_feedback_remaining.xlsx"):
    wb = openpyxl.Workbook()

    # sheet name update
    wb['Sheet'].title = "Sheet1"
    sh1 = wb.active
    # create iterable object- Here we are creating list of tuples
    data = [('rollno','register_sem','schedule_sem','subno','Name','email','aemail','contact')]
    # run for loop and append the record one by one
    for i in data:
        sh1.append(i)
    # save the workbook
    wb.save("course_feedback_remaining.xlsx")
feedback_not_submitted()