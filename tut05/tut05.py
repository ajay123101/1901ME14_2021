import csv
import os
from openpyxl import Workbook

roll_data = {}
name = {}

def generate_marksheet():
    if not os.path.isdir("output/"):
        os.mkdir("output/")
    with open("grades.csv", 'r') as roll_file:
        reader = csv.DictReader(roll_file)
        i = 1
        sem = 0
        dict = {}
        p = ['Sl No.','Subject No.', 'Subject Name', 'L-T-P', 'Credit', 'Subject Type', 'Grade']
        for row in reader:
            wb = Workbook()
            Roll = row['Roll']
            SubCode = row['SubCode']
            SubName = roll_data[SubCode][0]
            Ltp = roll_data[SubCode][1]
            Credit = row['Credit']
            Grade = row['Grade']
            Sem = row['Sem']
            Sub_Type = row['Sub_Type']

            if Roll not in dict:
                dict[Roll] = wb
                sem = 1
            if Sem != sem:
                i = 0
                sem = Sem
                ws1 = dict[Roll].create_sheet("Sheet_A")
                ws1.title = f"Sem{Sem}"
                ws1.append(p)
            dict[Roll][f'Sem{sem}'].append([i, SubCode, SubName, Ltp, int(Credit), Sub_Type, Grade])
            i += 1
        for key in dict:
            wb = dict[key]
            wb_new = wb['Sheet']
            wb_new.append(['Roll No.', key])
            wb_new.append(['Name of Student.', name[key]])
            wb_new.append(['Discipline', f"{key[4]}{key[5]}"])
            wb_new.append(['Semester No.', 1,2,3,4,5,6,7,8])
            c = []
            spi = []
            tct = []
            cpi = []
            sum_n = 0
            for_cpi = 0
            for i in range(1,9):
                n = wb[f'Sem{i}']
                sum = 0
                sum2 = 0
                for row in range(2, n.max_row+1):
                    sum += n.cell(column=5, row=row).value
                    if n.cell(column=7, row=row).value == 'AA':
                        sum2 += n.cell(column=5, row=row).value*10
                    elif n.cell(column=7, row=row).value == 'AB':
                        sum2 += n.cell(column=5, row=row).value*9
                    elif n.cell(column=7, row=row).value == 'BB':
                        sum2 += n.cell(column=5, row=row).value*8
                    elif n.cell(column=7, row=row).value == 'BC':
                        sum2 += n.cell(column=5, row=row).value*7
                    elif n.cell(column=7, row=row).value == 'CC':
                        sum2 += n.cell(column=5, row=row).value*6
                    elif n.cell(column=7, row=row).value == 'CD':
                        sum2 += n.cell(column=5, row=row).value*5
                    elif n.cell(column=7, row=row).value != 'DD':
                        sum2 += n.cell(column=5, row=row).value*4
                    elif n.cell(column=7, row=row).value == 'F':
                        sum2 += 0
                    elif n.cell(column=7, row=row).value == 'I':
                        sum2 += 0
                for_cpi += (sum2/sum)*sum
                sum_n += sum
                tct.append(sum_n)
                c.append(sum)
                spi.append(sum2/sum)
                cpi.append(for_cpi/sum_n)

            c.insert(0, 'Semester wise Credit Taken')
            spi.insert(0, 'SPI')
            tct.insert(0, 'Total Credits Taken')
            cpi.insert(0, 'CPI')

            wb_new.append(c)
            wb_new.append(spi)
            wb_new.append(tct)
            wb_new.append(cpi)

            wb_new.title = "Overall"





            wb.save('output/' + key + ".xlsx")
    return

with open('names-roll.csv', 'r') as temp:
    read = csv.DictReader(temp)
    for row in read:
        name[row['Roll']] = row['Name']


with open('subjects_master.csv', 'r') as temp:
    read = csv.DictReader(temp)
    for row in read:
        roll_data[row['subno']] = [row['subname'], row['ltp']]
    generate_marksheet()